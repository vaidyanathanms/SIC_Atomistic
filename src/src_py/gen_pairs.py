import math
from itertools import combinations_with_replacement
from openpyxl import load_workbook, Workbook

input_file = "../../atomtypes_charges/All_AtomandTopoTypes.xlsx"
input_sheet = "All_Unique_Types"
output_file = "../src_lmp/pair_coeff.dat"

def geom_mean(a, b):
    return round(math.sqrt(a * b), 5)

# Load input workbook
wb = load_workbook(input_file, data_only=True)
ws = wb[input_sheet]

# Read header row
header_row = 2
headers = [cell.value for cell in ws[header_row]]
header_map = {h: i for i, h in enumerate(headers) if h is not None}

required_cols = [
    "NewType_ID-1",
    "NewType_ID-2",
    "epsilon (kj/mol)",
    "sigma (Angstrom)",
    "cut-off rc (Angstrom)",
    "#Comments",
]

for col in required_cols:
    if col not in header_map:
        raise ValueError(f"Missing required column: {col}")

# Read all records
records = []
for row in ws.iter_rows(min_row=header_row+1, values_only=True):

    id_cell = row[header_map['NewType_ID-1']]
    if id_cell is None or str(id_cell).strip == "":
        break
    rec = {
        "id": int(row[header_map["NewType_ID-1"]]),
        "epsilon": float(row[header_map["epsilon (kj/mol)"]]),
        "sigma": float(row[header_map["sigma (Angstrom)"]]),
        "rc": float(row[header_map["cut-off rc (Angstrom)"]]),
        "label": str(row[header_map["#Comments"]]).replace("#", "").strip()[0:10],
    }
    records.append(rec)


# Write to output file
with open(output_file, "w") as f:

    f.write("# Pair Coefficients generated using geometric mixing\n")
    f.write(f'#NewType_ID-1\tNewType_ID-2\tepsilon (kcal/mol)\tsigma (Angstrom)'
            '\tcut-off rc (Angstrom)\t # Comments \n\n')
    
    for r1, r2 in combinations_with_replacement(records, 2):
        eps_ij = geom_mean(r1["epsilon"], r2["epsilon"])
        sig_ij = geom_mean(r1["sigma"], r2["sigma"])
        rc_ij  = geom_mean(r1["rc"], r2["rc"])

        label = f"{r1['label']}-{r2['label']}"

        line = (
            f"pair_coeff {r1['id']} {r2['id']} "
            f"{eps_ij:.5f} {sig_ij:.5f} {rc_ij:.5f} "
            f"# {label}\n"
        )

        f.write(line)

print(f"Done. Output written to: {output_file}")
